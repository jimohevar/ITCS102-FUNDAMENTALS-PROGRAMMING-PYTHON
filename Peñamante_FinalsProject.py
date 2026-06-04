import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op

try:
    workbook = op.load_workbook("Penamante_Database.xlsx")
except:
    workbook = op.Workbook()
    sheet = workbook.active

    sheet['A1'] = "ID"
    sheet['B1'] = "Patient Name"
    sheet['C1'] = "Age"
    sheet['D1'] = "Weight"
    sheet['E1'] = "Height"

    workbook.save("Penamante_Database.xlsx")

def validate_input():

    id = id_entry.get()
    name = name_entry.get()
    age = age_entry.get()
    weight = weight_entry.get()
    height = height_entry.get()

    if not id or not name or not age or not weight or not height:
        messagebox.showerror("Error","All are required.")
        return False

    if not age.isdigit():
        messagebox.showerror("Error","Age must be a number.")
        return False

    return True

def append_excel():

    if not validate_input():
        return

    workbook = op.load_workbook("Penamante_Database.xlsx")
    sheet = workbook.active

    sheet.append([id_entry.get(),name_entry.get(),age_entry.get(),weight_entry.get(),height_entry.get()])

    workbook.save("Penamante_Database.xlsx")

    messagebox.showinfo("Success","Record added successfully!")

    display_excel()

def display_excel():

    workbook = op.load_workbook("Penamante_Database.xlsx")
    sheet = workbook.active

    for row in tree.get_children():
        tree.delete(row)

    for row in sheet.iter_rows(
            min_row=2,
            values_only=True):

        tree.insert("",tk.END,values=row)

def update_data():

    selected = tree.focus()

    if not selected:
        messagebox.showerror("Error","Select a record first")
        return

    values = tree.item(selected,"values")

    record_id = values[0]

    workbook = op.load_workbook("Penamante_Database.xlsx")
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):

        if row[0].value == record_id:

            row[1].value = name_entry.get()
            row[2].value = age_entry.get()
            row[3].value = weight_entry.get()
            row[4].value = height_entry.get()

    workbook.save("Penamante_Database.xlsx")

    messagebox.showinfo("Success","Record updated successfully!")

    display_excel()

def delete_data():

    selected = tree.focus()

    if not selected:
        messagebox.showerror("Error","Select a record first")
        return

    values = tree.item(selected,"values")

    record_id = values[0]

    workbook = op.load_workbook("Penamante_Database.xlsx")
    sheet = workbook.active

    for i, row in enumerate(
            sheet.iter_rows(min_row=2),
            start=2):

        if row[0].value == record_id:
            sheet.delete_rows(i)
            break

    workbook.save("Penamante_Database.xlsx")

    messagebox.showinfo("Success","Record deleted successfully!")

    display_excel()

def select_record(event):

    selected = tree.focus()

    if not selected:
        return

    values = tree.item(selected,"values")

    id_entry.delete(0, tk.END)
    name_entry.delete(0, tk.END)
    age_entry.delete(0, tk.END)
    weight_entry.delete(0, tk.END)
    height_entry.delete(0, tk.END)

    id_entry.insert(0, values[0])
    name_entry.insert(0, values[1])
    age_entry.insert(0, values[2])
    weight_entry.insert(0, values[3])
    height_entry.insert(0, values[4])

root = tk.Tk()
root.title("Diet Monitoring System")
root.geometry("700x500")

tk.Label(root, text="ID").grid(row=0, column=0)
id_entry = tk.Entry(root)
id_entry.grid(row=0, column=1)

tk.Label(root, text="Patient Name").grid(row=1, column=0)
name_entry = tk.Entry(root)
name_entry.grid(row=1, column=1)

tk.Label(root, text="Age").grid(row=2, column=0)
age_entry = tk.Entry(root)
age_entry.grid(row=2, column=1)

tk.Label(root, text="Weight").grid(row=3, column=0)
weight_entry = tk.Entry(root)
weight_entry.grid(row=3, column=1)

tk.Label(root, text="Height").grid(row=4, column=0)
height_entry = tk.Entry(root)
height_entry.grid(row=4, column=1)

tk.Button(root,text="Add",command=append_excel).grid(row=5, column=0)

tk.Button(root,text="Update",command=update_data).grid(row=5, column=1)

tk.Button(root,text="Delete",command=delete_data).grid(row=5, column=2)

tree = ttk.Treeview(root,columns=("ID","Name","Age","Weight","Height"),show="headings")

tree.heading("ID", text="ID")
tree.heading("Name", text="Patient Name")
tree.heading("Age", text="Age")
tree.heading("Weight", text="Weight")
tree.heading("Height", text="Height")

tree.grid(row=6,column=0,columnspan=3,padx=10,pady=10)

tree.bind("<<TreeviewSelect>>", select_record)

display_excel()

root.mainloop()